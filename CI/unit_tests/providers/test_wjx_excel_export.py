from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from software.core.reverse_fill.schema import (
    REVERSE_FILL_FORMAT_AUTO,
    REVERSE_FILL_FORMAT_WJX_SCORE,
    REVERSE_FILL_FORMAT_WJX_SEQUENCE,
    REVERSE_FILL_FORMAT_WJX_TEXT,
)
from software.io.spreadsheets.wjx_excel import load_wjx_excel_export


def _write_workbook(path: Path, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


class WjxExcelExportTests:
    def test_load_wjx_excel_export_reads_question_columns_and_rows(self, tmp_path: Path) -> None:
        path = _write_workbook(
            tmp_path / "export.xlsx",
            [
                ["序号", "1、单选题", "2. 姓名", "备注"],
                [1, 2, "张三", "ignored"],
                [2, 1, "李四", "ignored"],
            ],
        )

        export = load_wjx_excel_export(str(path))

        assert export.source_path == str(path.resolve())
        assert export.header_row_number == 1
        assert export.total_data_rows == 2
        assert export.detected_format == REVERSE_FILL_FORMAT_WJX_SCORE
        assert export.selected_format == REVERSE_FILL_FORMAT_WJX_SCORE
        assert sorted(export.question_columns) == [1, 2]
        assert export.question_columns[1][0].column_index == 2
        assert export.question_columns[2][0].header == "2. 姓名"
        assert export.raw_rows[0].worksheet_row_number == 2
        assert export.raw_rows[0].values_by_column == {2: 2, 3: "张三"}

    def test_load_wjx_excel_export_detects_sequence_format_from_option_suffix(self, tmp_path: Path) -> None:
        path = _write_workbook(
            tmp_path / "sequence.xlsx",
            [["1、(选项1)", "1、(选项2)"], [1, 2]],
        )

        export = load_wjx_excel_export(str(path))

        assert export.detected_format == REVERSE_FILL_FORMAT_WJX_SEQUENCE

    def test_load_wjx_excel_export_detects_text_format_for_numeric_strings(self, tmp_path: Path) -> None:
        path = _write_workbook(tmp_path / "text.xlsx", [["1、满意度"], ["4"]])

        export = load_wjx_excel_export(str(path))

        assert export.detected_format == REVERSE_FILL_FORMAT_WJX_TEXT

    def test_load_wjx_excel_export_uses_preferred_format_when_given(self, tmp_path: Path) -> None:
        path = _write_workbook(tmp_path / "preferred.xlsx", [["1、满意度"], [5]])

        export = load_wjx_excel_export(str(path), preferred_format=REVERSE_FILL_FORMAT_WJX_TEXT)

        assert export.detected_format == REVERSE_FILL_FORMAT_WJX_SCORE
        assert export.selected_format == REVERSE_FILL_FORMAT_WJX_TEXT

    def test_load_wjx_excel_export_keeps_unknown_preferred_format_for_later_validation(self, tmp_path: Path) -> None:
        path = _write_workbook(tmp_path / "unknown.xlsx", [["1、满意度"], [5]])

        export = load_wjx_excel_export(str(path), preferred_format="custom_format")

        assert export.selected_format == "custom_format"

    def test_load_wjx_excel_export_rejects_missing_path(self) -> None:
        with pytest.raises(ValueError, match="未提供 Excel 文件路径"):
            load_wjx_excel_export("  ")

    def test_load_wjx_excel_export_rejects_nonexistent_file(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="Excel 文件不存在"):
            load_wjx_excel_export(str(tmp_path / "missing.xlsx"))

    def test_load_wjx_excel_export_rejects_empty_header_and_closes_workbook(self, tmp_path: Path) -> None:
        path = _write_workbook(tmp_path / "empty_header.xlsx", [])

        with pytest.raises(ValueError, match="Excel 缺少表头"):
            load_wjx_excel_export(str(path))

    def test_load_wjx_excel_export_closes_workbook_when_reader_raises(self, tmp_path: Path) -> None:
        path = tmp_path / "mocked.xlsx"
        path.write_bytes(b"not a real workbook")
        workbook = MagicMock()
        workbook.sheetnames = ["Sheet1"]
        worksheet = SimpleNamespace(iter_rows=MagicMock(side_effect=RuntimeError("reader failed")))
        workbook.__getitem__.return_value = worksheet

        with patch("os.path.exists", return_value=True), patch(
            "openpyxl.load_workbook",
            return_value=workbook,
        ):
            with pytest.raises(RuntimeError, match="reader failed"):
                load_wjx_excel_export(str(path), preferred_format=REVERSE_FILL_FORMAT_AUTO)

        workbook.close.assert_called_once_with()
